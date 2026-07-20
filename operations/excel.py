"""
Excel 操作模块 — 基于 openpyxl 的 Excel 读写与流程配置驱动
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List, Optional

class ExcelOperations:
    """Excel 操作器，支持打开工作簿、读取单元格/行/列/范围等操作"""

    def __init__(self):
        self.workbook = None
        self.file_path = None

    def open_workbook(self, file_path: str):
        """打开 Excel 文件"""
        self.file_path = file_path
        self.workbook = load_workbook(file_path, data_only=True)

    def get_sheet_names(self) -> List[str]:
        """获取所有工作表名称"""
        if self.workbook:
            return self.workbook.sheetnames
        return []

    def get_cell_value(self, sheet_name: str, cell_ref: str) -> Any:
        """获取指定单元格的值"""
        if not self.workbook:
            return None
        if sheet_name not in self.workbook.sheetnames:
            return None
        sheet = self.workbook[sheet_name]
        return sheet[cell_ref].value

    def get_row_values(self, sheet_name: str, row: int) -> List[Any]:
        """获取指定行的所有值"""
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return []
        sheet = self.workbook[sheet_name]
        return [cell.value for cell in sheet[row]]

    def get_column_values(self, sheet_name: str, column: int) -> List[Any]:
        """获取指定列的所有值"""
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return []
        sheet = self.workbook[sheet_name]
        col_letter = get_column_letter(column)
        return [cell.value for cell in sheet[col_letter]]

    def get_range_values(self, sheet_name: str, start_cell: str, end_cell: str) -> List[List[Any]]:
        """获取指定范围内的值"""
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return []
        sheet = self.workbook[sheet_name]
        return [[cell.value for cell in row] for row in sheet[start_cell:end_cell]]

    def get_sheet_data(self, sheet_name: str) -> Dict[str, Any]:
        """获取整个工作表数据，以字典形式返回"""
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return {}
        sheet = self.workbook[sheet_name]
        data = {}
        for row in sheet.iter_rows(values_only=True):
            for i, cell in enumerate(row):
                col_letter = get_column_letter(i + 1)
                data[f"{col_letter}{row[0]}"] = cell
        return data

    def close_workbook(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None

    def get_row_count(self, sheet_name: str) -> int:
        """获取行数"""
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return 0
        sheet = self.workbook[sheet_name]
        return sheet.max_row

    def get_column_count(self, sheet_name: str) -> int:
        """获取列数"""
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return 0
        sheet = self.workbook[sheet_name]
        return sheet.max_column

    def read_from_config(self, config: Dict[str, Any], variable_manager) -> Any:
        """从流程配置中读取 Excel 数据并存入变量"""
        file_path = variable_manager.resolve_expression(config.get("file_path", ""))
        sheet = config.get("sheet", "Sheet1")
        read_range = config.get("read_range", "cell")
        output_variable = config.get("output_variable", "")
        var_format = config.get("var_format", "string")

        variable_manager.load_excel(file_path)

        if read_range == "cell":
            cell_address = config.get("cell_address", "A1")
            value = variable_manager.get_excel_value(sheet, cell_address)
        elif read_range == "row":
            row_number = config.get("row_number", 1)
            value = variable_manager.get_excel_value(sheet, f"{row_number}:{row_number}")
        elif read_range == "column":
            column_number = config.get("column_number", 1)
            col_letter = get_column_letter(column_number)
            value = variable_manager.get_excel_value(sheet, f"{col_letter}:{col_letter}")
        elif read_range == "range":
            start_cell = config.get("start_cell", "A1")
            end_cell = config.get("end_cell", "A1")
            value = variable_manager.get_excel_value(sheet, f"{start_cell}:{end_cell}")
        else:
            return None

        if var_format == "number":
            try:
                value = float(value)
            except (ValueError, TypeError):
                pass
        elif var_format == "list":
            if not isinstance(value, list):
                value = [value]

        if output_variable:
            variable_manager.set_variable(output_variable, value)

        return value