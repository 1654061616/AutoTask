import pytest
import sys
import tempfile
import os
sys.path.insert(0, '..')
from core.engine import FlowEngine

def test_engine_initialization():
    engine = FlowEngine()
    assert not engine.is_running
    assert engine.current_step is None

def test_load_flow():
    engine = FlowEngine()
    flow = {
        "name": "Test",
        "version": "1.0",
        "steps": []
    }
    engine.load_flow(flow)
    assert engine.flow["name"] == "Test"

def test_stop_engine():
    engine = FlowEngine()
    engine.is_running = True
    engine.stop()
    assert not engine.is_running


class FakeKeyboard:
    def __init__(self):
        self.typed = []

    def type(self, text, interval=0):
        self.typed.append(text)


def test_keyboard_type_manual_backward_compat():
    engine = FlowEngine()
    executor = engine.step_executor
    kb = FakeKeyboard()
    executor.keyboard = kb
    config = {"input_text": "旧格式文本", "interval": 0.05}
    executor._execute_keyboard_type(config, {})
    assert kb.typed[0] == "旧格式文本"


def test_keyboard_type_variable():
    engine = FlowEngine()
    executor = engine.step_executor
    kb = FakeKeyboard()
    executor.keyboard = kb
    executor.variable_manager.set_variable("my_var", "变量内容")
    config = {"data_source": "variable", "variable_name": "my_var", "interval": 0.05}
    executor._execute_keyboard_type(config, {})
    assert kb.typed[0] == "变量内容"


def test_keyboard_type_excel_sequential():
    import openpyxl
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_path = tmp.name
    tmp.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws['A1'] = "第一行"
    ws['A2'] = "第二行"
    ws['A3'] = "第三行"
    wb.save(tmp_path)
    wb.close()

    try:
        engine = FlowEngine()
        executor = engine.step_executor
        kb = FakeKeyboard()
        executor.keyboard = kb
        excel_config = {
            "file_path": tmp_path, "sheet": "Sheet1",
            "read_mode": "sequential", "read_range": "cell",
            "cell_address": "A1", "var_format": "string",
        }
        config = {"data_source": "excel", "excel": excel_config, "interval": 0.05}
        excel_cursors = {}

        executor._execute_keyboard_type(config, excel_cursors)
        assert kb.typed[0] == "第一行"
        executor._execute_keyboard_type(config, excel_cursors)
        assert kb.typed[1] == "第二行"
        executor._execute_keyboard_type(config, excel_cursors)
        assert kb.typed[2] == "第三行"
        executor._execute_keyboard_type(config, excel_cursors)
        assert kb.typed[3] == "第一行"
    finally:
        os.unlink(tmp_path)


def test_keyboard_type_excel_random():
    import openpyxl
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_path = tmp.name
    tmp.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i in range(1, 11):
        ws.cell(row=i, column=1, value=f"行{i}")
    wb.save(tmp_path)
    wb.close()

    try:
        engine = FlowEngine()
        executor = engine.step_executor
        kb = FakeKeyboard()
        executor.keyboard = kb
        excel_config = {
            "file_path": tmp_path, "sheet": "Sheet1",
            "read_mode": "random", "read_range": "cell",
            "cell_address": "A1", "var_format": "string",
        }
        config = {"data_source": "excel", "excel": excel_config, "interval": 0.05}
        executor._execute_keyboard_type(config, {})
        assert kb.typed[0] in [f"行{i}" for i in range(1, 11)]
    finally:
        os.unlink(tmp_path)