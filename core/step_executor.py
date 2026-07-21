"""
步骤执行器 — 负责执行流程中的每一个步骤，分发到具体的操作模块
"""
import time
import random
from typing import Any, Callable, Dict


class StepExecutor:
    """步骤执行器，将步骤类型分发到对应的操作方法"""

    def __init__(self, mouse, keyboard, image_recognition, window,
                 condition, loop, excel, flow_control,
                 variable_manager, logger):
        self.mouse = mouse                       # 鼠标操作模块
        self.keyboard = keyboard                 # 键盘操作模块
        self.image_recognition = image_recognition  # 图像识别模块
        self.window = window                     # 窗口操作模块
        self.condition = condition               # 条件判断模块
        self.loop = loop                         # 循环控制模块
        self.excel = excel                       # Excel 操作模块
        self.flow_control = flow_control         # 流程控制模块（goto/label）
        self.variable_manager = variable_manager # 变量管理器
        self.logger = logger                     # 日志记录器
        self.last_click_pos = None               # 记录上一次点击位置
        self.last_move_pos = None                # 记录上一次移动位置

    def execute(self, step: Dict[str, Any], is_running_func: Callable[[], bool],
                excel_cursors: Dict[str, int], flow_nodes: list) -> Any:
        """根据步骤类型分发到对应的执行方法"""
        step_type = step["type"]
        config = step.get("config", {})
        step_result = None

        self.logger.info(f"执行操作类型: {step_type}")

        try:
            if step_type == "mouse_click":
                self._execute_mouse_click(config)
            elif step_type == "mouse_move":
                self._execute_mouse_move(config)
            elif step_type == "mouse_drag":
                self._execute_mouse_drag(config)
            elif step_type == "mouse_scroll":
                self._execute_mouse_scroll(config)
            elif step_type == "keyboard_type":
                self._execute_keyboard_type(config, excel_cursors)
            elif step_type == "keyboard_press":
                self._execute_keyboard_press(config)
            elif step_type == "keyboard_hotkey":
                self._execute_keyboard_hotkey(config)
            elif step_type == "wait":
                self._execute_wait(config)
            elif step_type == "image_find":
                step_result = self._execute_image_find(config, is_running_func)
            elif step_type == "image_click":
                step_result = self._execute_image_click(config, is_running_func)
            elif step_type == "image_exists":
                step_result = self._execute_image_exists(config, is_running_func)
            elif step_type == "window_find":
                self._execute_window_find(config)
            elif step_type == "window_activate":
                self._execute_window_activate(config)
            elif step_type == "window_close":
                self._execute_window_close(config)
            elif step_type == "log":
                self._execute_log(config)
            elif step_type == "if_else":
                step_result = self._execute_if_else(config)
            elif step_type == "loop":
                step_result = self._execute_loop(config, step["id"])
            elif step_type == "set_variable":
                self._execute_set_variable(config)
            elif step_type == "get_variable":
                self._execute_get_variable(config)
            elif step_type == "excel_read":
                self._execute_excel_read(config)
            elif step_type == "goto":
                self._execute_goto(config, flow_nodes)
            elif step_type == "label":
                pass
            else:
                self.logger.warning(f"未支持的步骤类型: {step_type}")
        except Exception as e:
            self.logger.error(f"执行步骤失败 [{step_type}]: {str(e)}")

        return step_result

    def _resolve_position(self, config):
        """解析坐标，支持绝对坐标和相对坐标"""
        position_type = config.get("position_type", "screen")
        if position_type == "relative":
            relative_base = config.get("relative_base", "current")
            rel_x = config.get("relative_x", 0)
            rel_y = config.get("relative_y", 0)
            if relative_base == "last_click" and self.last_click_pos:
                base_x, base_y = self.last_click_pos
            elif relative_base == "last_move" and self.last_move_pos:
                base_x, base_y = self.last_move_pos
            else:
                base_x, base_y = self.mouse.get_position()
            x = base_x + rel_x
            y = base_y + rel_y
            self.logger.info(f"相对坐标: 基准({relative_base})=({base_x}, {base_y}), 偏移=({rel_x}, {rel_y}), 目标=({x}, {y})")
            return x, y
        else:
            return config.get("x"), config.get("y")

    def _execute_mouse_click(self, config):
        """执行鼠标点击操作"""
        if not self.mouse:
            self.logger.error("鼠标操作模块未加载")
            return
        x, y = self._resolve_position(config)
        button = config.get("button", "left")
        clicks = config.get("clicks", 1)
        interval = config.get("interval", 0.2)
        # 随机偏移：在目标坐标周围随机偏移，避免被检测为脚本
        if config.get("random_offset", False) and x is not None and y is not None:
            offset_range = config.get("random_range", 5)
            offset_x = random.randint(-offset_range, offset_range)
            offset_y = random.randint(-offset_range, offset_range)
            x += offset_x
            y += offset_y
        if x is not None and y is not None:
            self.logger.info(f"鼠标点击: ({x}, {y}), 按钮: {button}, 次数: {clicks}")
            self.mouse.click(x=x, y=y, button=button, clicks=clicks, interval=interval)
            self.last_click_pos = (x, y)
        else:
            self.logger.info(f"鼠标点击: 当前位置, 按钮: {button}, 次数: {clicks}")
            self.mouse.click(button=button, clicks=clicks, interval=interval)
            self.last_click_pos = self.mouse.get_position()

    def _execute_mouse_move(self, config):
        """执行鼠标移动操作"""
        if not self.mouse:
            self.logger.error("鼠标操作模块未加载")
            return
        x, y = self._resolve_position(config)
        duration = config.get("duration", 0.5)
        move_type = config.get("move_type", "ease")
        if x is not None and y is not None:
            self.logger.info(f"鼠标移动: ({x}, {y}), 耗时: {duration}秒, 方式: {move_type}")
            if move_type == "linear":
                # 线性移动，不使用平滑曲线(匀速直线移动)
                self.mouse.move(x, y, duration=duration, easing=False)
            elif move_type == "random":
                offset_x = random.randint(-5, 5)
                offset_y = random.randint(-5, 5)
                self.mouse.move(x + offset_x, y + offset_y, duration=duration)
            else:
                # 使用默认的平滑曲线(贝塞尔曲线)
                self.mouse.move(x, y, duration=duration)
            self.last_move_pos = (x, y)

    def _execute_mouse_drag(self, config):
        """执行鼠标拖拽操作"""
        if not self.mouse:
            self.logger.error("鼠标操作模块未加载")
            return
        start_x = config.get("start_x")
        start_y = config.get("start_y")
        end_x = config.get("end_x")
        end_y = config.get("end_y")
        duration = config.get("duration", 0.5)
        if all([start_x, start_y, end_x, end_y]):
            self.logger.info(f"鼠标拖拽: ({start_x}, {start_y}) -> ({end_x}, {end_y})")
            self.mouse.drag(start_x, start_y, end_x, end_y, duration=duration)

    def _execute_mouse_scroll(self, config):
        """执行鼠标滚轮操作"""
        if not self.mouse:
            self.logger.error("鼠标操作模块未加载")
            return
        amount = config.get("amount", 100)
        x = config.get("x")
        y = config.get("y")
        self.logger.info(f"鼠标滚动: {amount}")
        self.mouse.scroll(amount, x, y)

    def _execute_keyboard_type(self, config, excel_cursors):
        """执行键盘输入操作，支持手动输入/Excel数据/变量三种数据源"""
        if not self.keyboard:
            self.logger.error("键盘操作模块未加载")
            return
        data_source = config.get("data_source", "manual")
        if data_source == "excel":
            text = self._read_excel_for_keyboard(config.get("excel", {}), excel_cursors)
        elif data_source == "variable":
            var_name = config.get("variable_name", "")
            text = str(self.variable_manager.get_variable(var_name) or "")
        else:
            text = config.get("input_text", "")
        text = self.variable_manager.resolve_expression(text)
        interval = config.get("interval", 0.05)
        self.logger.info(f"键盘输入: {text}")
        self.keyboard.type(text, interval=interval)

    def _read_excel_for_keyboard(self, excel_config, excel_cursors):
        """从 Excel 读取数据用于键盘输入"""
        import openpyxl
        file_path = excel_config.get("file_path", "")
        sheet = excel_config.get("sheet", "Sheet1")
        read_mode = excel_config.get("read_mode", "sequential")
        read_range = excel_config.get("read_range", "cell")
        var_format = excel_config.get("var_format", "string")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet]
        if read_mode == "sequential":
            cursor_key = f"{file_path}:{sheet}"
            current_row = excel_cursors.get(cursor_key, 1)
            if current_row > ws.max_row:
                current_row = 1
            value = self._read_excel_row_value(ws, current_row, read_range, excel_config)
            excel_cursors[cursor_key] = current_row + 1
        else:
            current_row = random.randint(1, ws.max_row)
            value = self._read_excel_row_value(ws, current_row, read_range, excel_config)
        wb.close()
        return self._format_excel_value(value, var_format)

    def _read_excel_row_value(self, ws, row_num, read_range, config):
        """根据读取模式从 Excel 行中获取值"""
        if read_range == "cell":
            addr = config.get("cell_address", "A1")
            col_letter = ''.join(c for c in addr if c.isalpha())
            return ws[f"{col_letter}{row_num}"].value
        elif read_range == "row":
            return [ws.cell(row=row_num, column=c).value for c in range(1, ws.max_column + 1)]
        elif read_range == "column":
            col_num = config.get("column_number", 1)
            return ws.cell(row=row_num, column=col_num).value
        elif read_range == "range":
            import openpyxl.utils
            start_cell = config.get("start_cell", "A1")
            end_cell = config.get("end_cell", "A1")
            start_col = openpyxl.utils.column_index_from_string(
                ''.join(c for c in start_cell if c.isalpha()))
            start_row = int(''.join(c for c in start_cell if c.isdigit()))
            end_col = openpyxl.utils.column_index_from_string(
                ''.join(c for c in end_cell if c.isalpha()))
            end_row = int(''.join(c for c in end_cell if c.isdigit()))
            total_rows = end_row - start_row + 1
            offset = (row_num - 1) % total_rows
            actual_row = start_row + offset
            return [ws.cell(row=actual_row, column=c).value for c in range(start_col, end_col + 1)]
        return None

    def _format_excel_value(self, value, var_format):
        """根据格式类型格式化 Excel 值"""
        if var_format == "number":
            return str(value) if value is not None else "0"
        elif var_format == "list":
            if isinstance(value, list):
                return ", ".join(str(v) if v is not None else "" for v in value)
            return str(value) if value is not None else ""
        else:
            return str(value) if value is not None else ""

    def _execute_keyboard_press(self, config):
        """执行按键操作"""
        if not self.keyboard:
            self.logger.error("键盘操作模块未加载")
            return
        key = config.get("key", "")
        presses = config.get("presses", 1)
        self.logger.info(f"按键按下: {key}, 次数: {presses}")
        self.keyboard.press(key, presses=presses)

    def _execute_keyboard_hotkey(self, config):
        """执行快捷键组合操作"""
        if not self.keyboard:
            self.logger.error("键盘操作模块未加载")
            return
        keys = config.get("keys", [])
        self.logger.info(f"快捷键: {keys}")
        self.keyboard.hotkey(*keys)

    def _execute_image_find(self, config, is_running_func):
        """执行图像查找，支持等待模式"""
        if not self.image_recognition:
            self.logger.error("图像识别模块未加载")
            return False
        image_path = config.get("image_path", "")
        region = config.get("region")
        similarity = config.get("similarity", 0.8)
        algorithm = config.get("algorithm", "template")
        direction = config.get("direction", "default")
        wait_find = config.get("wait_find", False)
        wait_timeout = config.get("wait_timeout", 5)
        self.logger.info(f"查找图片: {image_path}, 相似度: {similarity}, 算法: {algorithm}, 方向: {direction}, 超时: {wait_timeout}秒")
        start_time = time.time()
        result = None
        while True:
            result = self.image_recognition.find_image(image_path, region=region, threshold=similarity, method=algorithm, direction=direction)
            if result:
                self.logger.info(f"找到图片: {result}")
                self.variable_manager.set_variable("image_result", result)
                return True
            if not wait_find:
                break
            elapsed_time = time.time() - start_time
            if elapsed_time >= wait_timeout:
                self.logger.info(f"查找图片超时({wait_timeout}秒)")
                break
            if not is_running_func():
                self.logger.info("用户停止，中断查找图片")
                break
            time.sleep(0.5)
        self.logger.info("未找到图片")
        self.variable_manager.set_variable("image_result", None)
        return False

    def _execute_image_click(self, config, is_running_func):
        """执行图像点击，找到图片后点击其位置"""
        if not self.image_recognition or not self.mouse:
            self.logger.error("图像识别或鼠标模块未加载")
            return False
        image_path = config.get("image_path", "")
        offset_x = config.get("offset_x", 0)
        offset_y = config.get("offset_y", 0)
        similarity = config.get("similarity", 0.8)
        algorithm = config.get("algorithm", "template")
        direction = config.get("direction", "default")
        wait_find = config.get("wait_find", False)
        wait_timeout = config.get("wait_timeout", 5)
        self.logger.info(f"点击图片: {image_path}, 相似度: {similarity}, 算法: {algorithm}, 方向: {direction}, 超时: {wait_timeout}秒")
        start_time = time.time()
        result = None
        while True:
            result = self.image_recognition.find_image(image_path, threshold=similarity, method=algorithm, direction=direction)
            if result:
                x, y = result
                self.mouse.click(x + offset_x, y + offset_y)
                self.logger.info(f"点击位置: ({x + offset_x}, {y + offset_y})")
                return True
            if not wait_find:
                break
            elapsed_time = time.time() - start_time
            if elapsed_time >= wait_timeout:
                self.logger.info(f"点击图片超时({wait_timeout}秒)")
                break
            if not is_running_func():
                self.logger.info("用户停止，中断点击图片")
                break
            time.sleep(0.5)
        self.logger.info("未找到图片，跳过点击")
        return False

    def _execute_image_exists(self, config, is_running_func):
        """判断图片是否存在，支持等待模式"""
        if not self.image_recognition:
            self.logger.error("图像识别模块未加载")
            return False
        image_path = config.get("image_path", "")
        similarity = config.get("similarity", 0.8)
        algorithm = config.get("algorithm", "template")
        direction = config.get("direction", "default")
        wait_find = config.get("wait_find", False)
        wait_timeout = config.get("wait_timeout", 5)
        self.logger.info(f"判断图片是否存在: {image_path}, 相似度: {similarity}, 算法: {algorithm}, 方向: {direction}, 超时: {wait_timeout}秒")
        start_time = time.time()
        exists = False
        while True:
            exists = self.image_recognition.image_exists(image_path, threshold=similarity, method=algorithm, direction=direction)
            if exists:
                self.logger.info("图片存在")
                self.variable_manager.set_variable("image_exists", True)
                return True
            if not wait_find:
                break
            elapsed_time = time.time() - start_time
            if elapsed_time >= wait_timeout:
                self.logger.info(f"判断图片存在超时({wait_timeout}秒)")
                break
            if not is_running_func():
                self.logger.info("用户停止，中断判断图片存在")
                break
            time.sleep(0.5)
        self.logger.info("图片不存在")
        self.variable_manager.set_variable("image_exists", False)
        return False

    def _execute_window_find(self, config):
        """查找窗口并将句柄存入变量"""
        if not self.window:
            self.logger.error("窗口操作模块未加载")
            return
        title = config.get("title", "")
        self.logger.info(f"查找窗口: {title}")
        handle = self.window.find_window(title)
        if handle:
            self.logger.info(f"找到窗口: {handle}")
            self.variable_manager.set_variable("window_handle", handle)
        else:
            self.logger.info("未找到窗口")

    def _execute_window_activate(self, config):
        """激活（置顶）指定窗口"""
        if not self.window:
            self.logger.error("窗口操作模块未加载")
            return
        title = config.get("title", "")
        handle = config.get("handle")
        if not handle and title:
            handle = self.window.find_window(title)
        if handle:
            self.logger.info(f"激活窗口: {handle}")
            self.window.activate_window(handle)
        else:
            self.logger.warning("未找到要激活的窗口")

    def _execute_window_close(self, config):
        """关闭指定窗口"""
        if not self.window:
            self.logger.error("窗口操作模块未加载")
            return
        title = config.get("title", "")
        handle = config.get("handle")
        if not handle and title:
            handle = self.window.find_window(title)
        if handle:
            self.logger.info(f"关闭窗口: {handle}")
            self.window.close_window(handle)
        else:
            self.logger.warning("未找到要关闭的窗口")

    def _execute_log(self, config):
        """输出日志信息"""
        message = config.get("message", "")
        level = config.get("level", "info")
        if level == "error":
            self.logger.error(message)
        elif level == "warning":
            self.logger.warning(message)
        else:
            self.logger.info(message)

    def _execute_if_else(self, config):
        """执行条件判断，返回 True/False"""
        return self.condition.evaluate_from_config(config, self.variable_manager)

    def _execute_loop(self, config, step_id):
        """执行循环控制"""
        return self.loop.evaluate(config, step_id, self.variable_manager)

    def _execute_set_variable(self, config):
        """设置变量值"""
        name = config.get("name", "")
        value = self.variable_manager.resolve_expression(str(config.get("value", "")))
        self.variable_manager.set_variable(name, value)

    def _execute_get_variable(self, config):
        """获取变量值并打印日志"""
        name = config.get("name", "")
        value = self.variable_manager.get_variable(name)
        self.logger.info(f"获取变量: {name} = {value}")

    def _execute_excel_read(self, config):
        """从 Excel 读取数据"""
        self.excel.read_from_config(config, self.variable_manager)

    def _execute_goto(self, config, flow_nodes):
        """执行跳转操作"""
        self.flow_control.goto(config, flow_nodes)

    def _execute_wait(self, wait_config):
        """等待指定时间，支持固定时长和随机时长"""
        if isinstance(wait_config, dict):
            wait_type = wait_config.get("type", "fixed")
            if wait_type == "random":
                min_wait = wait_config.get("min", 1)
                max_wait = wait_config.get("max", 3)
                wait_time = random.uniform(min_wait, max_wait)
            else:
                wait_time = wait_config.get("value", 1)
        else:
            wait_time = wait_config
        time.sleep(wait_time)


def create_step_executor(variable_manager, logger):
    """工厂函数：创建 StepExecutor 实例并注入所有操作模块"""
    from operations.mouse import MouseOperations
    from operations.keyboard import KeyboardOperations
    from operations.image_recognition import ImageRecognition
    from operations.window import WindowOperations
    from operations.condition import ConditionEvaluator
    from operations.loop import LoopController
    from operations.excel import ExcelOperations
    from operations.control import FlowControl

    mouse = MouseOperations()
    keyboard = KeyboardOperations()
    image_recognition = ImageRecognition()
    window = WindowOperations()
    condition = ConditionEvaluator()
    condition.set_image_ops(image_recognition)
    condition.set_window_ops(window)
    loop = LoopController()
    excel = ExcelOperations()
    flow_control = FlowControl()

    return StepExecutor(
        mouse=mouse,
        keyboard=keyboard,
        image_recognition=image_recognition,
        window=window,
        condition=condition,
        loop=loop,
        excel=excel,
        flow_control=flow_control,
        variable_manager=variable_manager,
        logger=logger,
    )