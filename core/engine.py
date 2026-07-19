from typing import Dict, Any, Optional, List, Callable
import threading
import time
from .parser import FlowParser
from .variables import VariableManager
from .logger import Logger

class FlowEngine:
    def __init__(self):
        self.flow: Optional[Dict[str, Any]] = None
        self.variable_manager = VariableManager()
        self.logger = Logger()
        self.is_running = False
        self.current_step = None
        self.thread: Optional[threading.Thread] = None
        self.parser = FlowParser()
        self._init_operations()
        self.completed_callbacks: List[Callable[[bool, str], None]] = []
        self._excel_cursors: Dict[str, int] = {}
    
    def add_completed_callback(self, callback: Callable[[bool, str], None]):
        """添加任务完成回调函数，当任务完成或异常时调用"""
        if callback not in self.completed_callbacks:
            self.completed_callbacks.append(callback)
    
    def remove_completed_callback(self, callback: Callable[[bool, str], None]):
        """移除任务完成回调函数"""
        if callback in self.completed_callbacks:
            self.completed_callbacks.remove(callback)
    
    def _init_operations(self):
        """初始化操作模块"""
        try:
            from operations.mouse import MouseOperations
            from operations.keyboard import KeyboardOperations
            from operations.image_recognition import ImageRecognition
            from operations.window import WindowOperations
            from operations.condition import ConditionEvaluator
            from operations.loop import LoopController
            from operations.excel import ExcelOperations
            from operations.control import FlowControl

            self.mouse = MouseOperations()
            self.keyboard = KeyboardOperations()
            self.image_recognition = ImageRecognition()
            self.window = WindowOperations()
            self.condition = ConditionEvaluator()
            self.condition.set_image_ops(self.image_recognition)
            self.condition.set_window_ops(self.window)
            self.loop = LoopController()
            self.excel = ExcelOperations()
            self.flow_control = FlowControl()
        except ImportError as e:
            self.logger.error(f"加载操作模块失败: {str(e)}")
            self.mouse = None
            self.keyboard = None
            self.image_recognition = None
            self.window = None
            self.condition = None
            self.loop = None
            self.excel = None
            self.flow_control = None
    
    def load_flow(self, flow: Dict[str, Any]):
        self.flow = self.parser.parse(flow)
    
    def load_flow_from_file(self, file_path: str):
        self.flow = self.parser.load_from_file(file_path)
    
    def set_excel_data(self, excel_path: str):
        self.variable_manager.load_excel(excel_path)
    
    def run(self):
        if not self.flow or self.is_running:
            return
        self.is_running = True
        self.thread = threading.Thread(target=self._execute_flow)
        self.thread.start()
    
    def _execute_flow(self):
        success = True
        error_message = ""
        try:
            self.logger.info("开始执行流程: {}".format(self.flow.get("name", "Unknown")))
            steps = self.flow.get("steps", [])
            if not steps:
                self.logger.warning("流程中没有步骤")
                return
            
            current_index = 0
            while current_index < len(steps) and self.is_running:
                step = steps[current_index]
                self.current_step = step
                self.logger.info(f"执行步骤: {step.get('name', step.get('id'))}")
                
                step_result = None
                try:
                    step_result = self._execute_step(step)
                except Exception as e:
                    self.logger.error(f"步骤执行失败: {str(e)}")
                    success = False
                    error_message = str(e)

                if self.flow_control and self.flow_control.goto_target is not None:
                    current_index = self.flow_control.goto_target
                    self.flow_control.clear_goto()
                    continue

                # 判断是否需要走分支路径
                true_step_id = step.get("true_step")
                false_step_id = step.get("false_step")
                
                if true_step_id is not None or false_step_id is not None:
                    # 分支节点，根据执行结果决定走哪个分支
                    if step_result:
                        self.logger.info(f"执行结果为True，走True分支")
                        if true_step_id:
                            current_index = self._find_step_index(true_step_id)
                        else:
                            self.logger.info("True分支连接到end节点，流程结束")
                            current_index = len(steps)
                    else:
                        self.logger.info(f"执行结果为False，走False分支")
                        if false_step_id:
                            current_index = self._find_step_index(false_step_id)
                        else:
                            self.logger.info("False分支连接到end节点，流程结束")
                            current_index = len(steps)
                else:
                    # 线性执行
                    next_step_id = step.get("next_step")
                    if next_step_id:
                        current_index = self._find_step_index(next_step_id)
                    else:
                        current_index += 1
            
            self.logger.info("流程执行完成")
        except Exception as e:
            self.logger.error(f"流程执行异常: {str(e)}")
            success = False
            error_message = str(e)
        finally:
            self.is_running = False
            self.current_step = None
            for callback in self.completed_callbacks:
                try:
                    callback(success, error_message)
                except Exception:
                    pass
    
    def _execute_step(self, step: Dict[str, Any]):
        """
        执行单个步骤方法: 根据步骤类型调用对应的操作模块
        
        参数:
            step: 步骤字典，包含type和config字段
        
        返回:
            step_result: 步骤执行结果，对于分支节点返回True/False，其他节点返回None
        """
        step_type = step["type"]
        config = step.get("config", {})
        step_result = None
        
        # 执行步骤前等待
        wait_before = step.get("wait_before")
        if wait_before:
            self._execute_wait(wait_before)
        
        self.logger.info(f"执行操作类型: {step_type}")
        
        # 根据步骤类型调用对应的操作
        try:
            if step_type == "mouse_click":
                self._execute_mouse_click(config)
            elif step_type == "mouse_move":
                self._execute_mouse_move(config)
            elif step_type == "mouse_drag":
                self._execute_mouse_drag(config)
            elif step_type == "mouse_scroll":
                self._execute_mouse_scroll(config)
            elif step_type == "keyboard_type":
                self._execute_keyboard_type(config)
            elif step_type == "keyboard_press":
                self._execute_keyboard_press(config)
            elif step_type == "keyboard_hotkey":
                self._execute_keyboard_hotkey(config)
            elif step_type == "wait":
                self._execute_wait(config)
            elif step_type == "image_find":
                step_result = self._execute_image_find(config)
            elif step_type == "image_click":
                step_result = self._execute_image_click(config)
            elif step_type == "image_exists":
                step_result = self._execute_image_exists(config)
            elif step_type == "window_find":
                self._execute_window_find(config)
            elif step_type == "window_activate":
                self._execute_window_activate(config)
            elif step_type == "window_close":
                self._execute_window_close(config)
            elif step_type == "log":
                self._execute_log(config)
            elif step_type == "if_else":
                step_result = self._execute_if_else(config)
            elif step_type == "loop":
                step_result = self._execute_loop(config, step["id"])
            elif step_type == "set_variable":
                self._execute_set_variable(config)
            elif step_type == "get_variable":
                self._execute_get_variable(config)
            elif step_type == "excel_read":
                self._execute_excel_read(config)
            elif step_type == "goto":
                self._execute_goto(config)
            elif step_type == "label":
                pass
            else:
                self.logger.warning(f"未支持的步骤类型: {step_type}")
        except Exception as e:
            self.logger.error(f"执行步骤失败 [{step_type}]: {str(e)}")
        
        # 执行步骤后等待
        wait_after = step.get("wait_after")
        if wait_after:
            self._execute_wait(wait_after)
        
        return step_result
    
    def _execute_mouse_click(self, config):
        """执行鼠标点击操作"""
        if not self.mouse:
            self.logger.error("鼠标操作模块未加载")
            return
        
        x = config.get("x")
        y = config.get("y")
        button = config.get("button", "left")
        clicks = config.get("clicks", 1)
        interval = config.get("interval", 0.2)
        
        if x is not None and y is not None:
            self.logger.info(f"鼠标点击: ({x}, {y}), 按钮: {button}, 次数: {clicks}")
            self.mouse.click(x=x, y=y, button=button, clicks=clicks, interval=interval)
        else:
            self.logger.info(f"鼠标点击: 当前位置, 按钮: {button}, 次数: {clicks}")
            self.mouse.click(button=button, clicks=clicks, interval=interval)
    
    def _execute_mouse_move(self, config):
        """执行鼠标移动操作"""
        if not self.mouse:
            self.logger.error("鼠标操作模块未加载")
            return
        
        x = config.get("x")
        y = config.get("y")
        duration = config.get("duration", 0.5)
        
        if x is not None and y is not None:
            self.logger.info(f"鼠标移动: ({x}, {y}), 耗时: {duration}秒")
            self.mouse.move(x, y, duration=duration)
    
    def _execute_mouse_drag(self, config):
        """执行鼠标拖拽操作"""
        if not self.mouse:
            self.logger.error("鼠标操作模块未加载")
            return
        
        start_x = config.get("start_x")
        start_y = config.get("start_y")
        end_x = config.get("end_x")
        end_y = config.get("end_y")
        duration = config.get("duration", 0.5)
        
        if all([start_x, start_y, end_x, end_y]):
            self.logger.info(f"鼠标拖拽: ({start_x}, {start_y}) -> ({end_x}, {end_y})")
            self.mouse.drag(start_x, start_y, end_x, end_y, duration=duration)
    
    def _execute_mouse_scroll(self, config):
        """执行鼠标滚动操作"""
        if not self.mouse:
            self.logger.error("鼠标操作模块未加载")
            return
        
        amount = config.get("amount", 100)
        x = config.get("x")
        y = config.get("y")
        
        self.logger.info(f"鼠标滚动: {amount}")
        self.mouse.scroll(amount, x, y)
    
    def _execute_keyboard_type(self, config):
        if not self.keyboard:
            self.logger.error("键盘操作模块未加载")
            return

        data_source = config.get("data_source", "manual")

        if data_source == "excel":
            text = self._read_excel_for_keyboard(config.get("excel", {}))
        elif data_source == "variable":
            var_name = config.get("variable_name", "")
            text = str(self.variable_manager.get_variable(var_name) or "")
        else:
            text = config.get("input_text", "")

        text = self.variable_manager.resolve_expression(text)
        interval = config.get("interval", 0.05)

        self.logger.info(f"键盘输入: {text}")
        self.keyboard.type(text, interval=interval)

    def _read_excel_for_keyboard(self, excel_config):
        import openpyxl
        import random

        file_path = excel_config.get("file_path", "")
        sheet = excel_config.get("sheet", "Sheet1")
        read_mode = excel_config.get("read_mode", "sequential")
        read_range = excel_config.get("read_range", "cell")
        var_format = excel_config.get("var_format", "string")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet]

        if read_mode == "sequential":
            cursor_key = f"{file_path}:{sheet}"
            current_row = self._excel_cursors.get(cursor_key, 1)
            if current_row > ws.max_row:
                current_row = 1
            value = self._read_excel_row_value(ws, current_row, read_range, excel_config)
            self._excel_cursors[cursor_key] = current_row + 1
        else:
            current_row = random.randint(1, ws.max_row)
            value = self._read_excel_row_value(ws, current_row, read_range, excel_config)

        wb.close()
        return self._format_excel_value(value, var_format)

    def _read_excel_row_value(self, ws, row_num, read_range, config):
        if read_range == "cell":
            addr = config.get("cell_address", "A1")
            col_letter = ''.join(c for c in addr if c.isalpha())
            return ws[f"{col_letter}{row_num}"].value
        elif read_range == "row":
            return [ws.cell(row=row_num, column=c).value for c in range(1, ws.max_column + 1)]
        elif read_range == "column":
            col_num = config.get("column_number", 1)
            return ws.cell(row=row_num, column=col_num).value
        elif read_range == "range":
            import openpyxl.utils
            start_cell = config.get("start_cell", "A1")
            end_cell = config.get("end_cell", "A1")
            start_col = openpyxl.utils.column_index_from_string(
                ''.join(c for c in start_cell if c.isalpha()))
            start_row = int(''.join(c for c in start_cell if c.isdigit()))
            end_col = openpyxl.utils.column_index_from_string(
                ''.join(c for c in end_cell if c.isalpha()))
            end_row = int(''.join(c for c in end_cell if c.isdigit()))
            total_rows = end_row - start_row + 1
            offset = (row_num - 1) % total_rows
            actual_row = start_row + offset
            return [ws.cell(row=actual_row, column=c).value for c in range(start_col, end_col + 1)]
        return None

    def _format_excel_value(self, value, var_format):
        if var_format == "number":
            return str(value) if value is not None else "0"
        elif var_format == "list":
            if isinstance(value, list):
                return ", ".join(str(v) if v is not None else "" for v in value)
            return str(value) if value is not None else ""
        else:
            return str(value) if value is not None else ""
    
    def _execute_keyboard_press(self, config):
        """执行按键按下操作"""
        if not self.keyboard:
            self.logger.error("键盘操作模块未加载")
            return
        
        key = config.get("key", "")
        presses = config.get("presses", 1)
        
        self.logger.info(f"按键按下: {key}, 次数: {presses}")
        self.keyboard.press(key, presses=presses)
    
    def _execute_keyboard_hotkey(self, config):
        """执行快捷键操作"""
        if not self.keyboard:
            self.logger.error("键盘操作模块未加载")
            return
        
        keys = config.get("keys", [])
        
        self.logger.info(f"快捷键: {keys}")
        self.keyboard.hotkey(*keys)
    
    def _execute_image_find(self, config):
        """执行图片查找操作"""
        if not self.image_recognition:
            self.logger.error("图像识别模块未加载")
            return False
        
        image_path = config.get("image_path", "")
        region = config.get("region")
        similarity = config.get("similarity", 0.8)
        algorithm = config.get("algorithm", "template")
        direction = config.get("direction", "default")
        wait_find = config.get("wait_find", False)
        wait_timeout = config.get("wait_timeout", 5)
        
        self.logger.info(f"查找图片: {image_path}, 相似度: {similarity}, 算法: {algorithm}, 方向: {direction}, 超时: {wait_timeout}秒")
        
        start_time = time.time()
        result = None
        
        while True:
            result = self.image_recognition.find_image(image_path, region=region, threshold=similarity, method=algorithm, direction=direction)
            if result:
                self.logger.info(f"找到图片: {result}")
                self.variable_manager.set_variable("image_result", result)
                return True
            
            if not wait_find:
                break
            
            elapsed_time = time.time() - start_time
            if elapsed_time >= wait_timeout:
                self.logger.info(f"查找图片超时({wait_timeout}秒)")
                break
            
            if not self.is_running:
                self.logger.info("用户停止，中断查找图片")
                break
            
            time.sleep(0.5)
        
        self.logger.info("未找到图片")
        self.variable_manager.set_variable("image_result", None)
        return False
    
    def _execute_image_click(self, config):
        """执行点击图片操作"""
        if not self.image_recognition or not self.mouse:
            self.logger.error("图像识别或鼠标模块未加载")
            return False
        
        image_path = config.get("image_path", "")
        offset_x = config.get("offset_x", 0)
        offset_y = config.get("offset_y", 0)
        similarity = config.get("similarity", 0.8)
        algorithm = config.get("algorithm", "template")
        direction = config.get("direction", "default")
        wait_find = config.get("wait_find", False)
        wait_timeout = config.get("wait_timeout", 5)
        
        self.logger.info(f"点击图片: {image_path}, 相似度: {similarity}, 算法: {algorithm}, 方向: {direction}, 超时: {wait_timeout}秒")
        
        start_time = time.time()
        result = None
        
        while True:
            result = self.image_recognition.find_image(image_path, threshold=similarity, method=algorithm, direction=direction)
            if result:
                x, y = result
                self.mouse.click(x + offset_x, y + offset_y)
                self.logger.info(f"点击位置: ({x + offset_x}, {y + offset_y})")
                return True
            
            if not wait_find:
                break
            
            elapsed_time = time.time() - start_time
            if elapsed_time >= wait_timeout:
                self.logger.info(f"点击图片超时({wait_timeout}秒)")
                break
            
            if not self.is_running:
                self.logger.info("用户停止，中断点击图片")
                break
            
            time.sleep(0.5)
        
        self.logger.info("未找到图片，跳过点击")
        return False
    
    def _execute_image_exists(self, config):
        """执行图片存在判断操作"""
        if not self.image_recognition:
            self.logger.error("图像识别模块未加载")
            return False
        
        image_path = config.get("image_path", "")
        similarity = config.get("similarity", 0.8)
        algorithm = config.get("algorithm", "template")
        direction = config.get("direction", "default")
        wait_find = config.get("wait_find", False)
        wait_timeout = config.get("wait_timeout", 5)
        
        self.logger.info(f"判断图片是否存在: {image_path}, 相似度: {similarity}, 算法: {algorithm}, 方向: {direction}, 超时: {wait_timeout}秒")
        
        start_time = time.time()
        exists = False
        
        while True:
            exists = self.image_recognition.image_exists(image_path, threshold=similarity, method=algorithm, direction=direction)
            if exists:
                self.logger.info("图片存在")
                self.variable_manager.set_variable("image_exists", True)
                return True
            
            if not wait_find:
                break
            
            elapsed_time = time.time() - start_time
            if elapsed_time >= wait_timeout:
                self.logger.info(f"判断图片存在超时({wait_timeout}秒)")
                break
            
            if not self.is_running:
                self.logger.info("用户停止，中断判断图片存在")
                break
            
            time.sleep(0.5)
        
        self.logger.info("图片不存在")
        self.variable_manager.set_variable("image_exists", False)
        return False
    
    def _execute_window_find(self, config):
        """执行窗口查找操作"""
        if not self.window:
            self.logger.error("窗口操作模块未加载")
            return
        
        title = config.get("title", "")
        
        self.logger.info(f"查找窗口: {title}")
        handle = self.window.find_window(title)
        if handle:
            self.logger.info(f"找到窗口: {handle}")
            self.variable_manager.set_variable("window_handle", handle)
        else:
            self.logger.info("未找到窗口")
    
    def _execute_window_activate(self, config):
        """执行激活窗口操作"""
        if not self.window:
            self.logger.error("窗口操作模块未加载")
            return
        
        title = config.get("title", "")
        handle = config.get("handle")
        
        # 如果没有提供handle，根据title查找窗口
        if not handle and title:
            handle = self.window.find_window(title)
        
        if handle:
            self.logger.info(f"激活窗口: {handle}")
            self.window.activate_window(handle)
        else:
            self.logger.warning("未找到要激活的窗口")
    
    def _execute_window_close(self, config):
        """执行关闭窗口操作"""
        if not self.window:
            self.logger.error("窗口操作模块未加载")
            return
        
        title = config.get("title", "")
        handle = config.get("handle")
        
        # 如果没有提供handle，根据title查找窗口
        if not handle and title:
            handle = self.window.find_window(title)
        
        if handle:
            self.logger.info(f"关闭窗口: {handle}")
            self.window.close_window(handle)
        else:
            self.logger.warning("未找到要关闭的窗口")
    
    def _execute_log(self, config):
        """执行日志操作"""
        message = config.get("message", "")
        level = config.get("level", "info")
        
        if level == "error":
            self.logger.error(message)
        elif level == "warning":
            self.logger.warning(message)
        else:
            self.logger.info(message)
    
    def _execute_if_else(self, config):
        return self.condition.evaluate_from_config(config, self.variable_manager)

    def _execute_loop(self, config, step_id):
        return self.loop.evaluate(config, step_id, self.variable_manager)

    def _execute_set_variable(self, config):
        name = config.get("name", "")
        value = self.variable_manager.resolve_expression(str(config.get("value", "")))
        self.variable_manager.set_variable(name, value)

    def _execute_get_variable(self, config):
        name = config.get("name", "")
        value = self.variable_manager.get_variable(name)
        self.logger.info(f"获取变量: {name} = {value}")

    def _execute_excel_read(self, config):
        self.excel.read_from_config(config, self.variable_manager)

    def _execute_goto(self, config):
        steps = self.flow.get("steps", [])
        self.flow_control.goto(config, steps)

    def _execute_wait(self, wait_config):
        if isinstance(wait_config, dict):
            wait_type = wait_config.get("type", "fixed")
            if wait_type == "random":
                min_wait = wait_config.get("min", 1)
                max_wait = wait_config.get("max", 3)
                import random
                wait_time = random.uniform(min_wait, max_wait)
            else:
                wait_time = wait_config.get("value", 1)
        else:
            wait_time = wait_config
        time.sleep(wait_time)
    
    def _find_step_index(self, step_id: str) -> int:
        steps = self.flow.get("steps", [])
        for i, step in enumerate(steps):
            if step["id"] == step_id:
                return i
        return -1
    
    def stop(self):
        self.is_running = False
        if self.thread:
            self.thread.join(timeout=5)
    
    def get_status(self) -> Dict[str, Any]:
        return {
            "is_running": self.is_running,
            "current_step": self.current_step,
            "flow_name": self.flow.get("name") if self.flow else None
        }